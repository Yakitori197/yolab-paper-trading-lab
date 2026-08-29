"""E7/E8 trade-ledger export: full regeneration (idempotent) of
data/trade_ledger.csv and data/trade_report.xlsx from data/paper.db.

E8 retires data/trade_summary.txt in favor of the "總覽" sheet in the xlsx
workbook -- write_summary()/build_summary()/format_section() are kept
in this module unchanged (zero-deletion) but main() no longer calls them,
and any pre-existing trade_summary.txt on disk is left alone.

Read-only: opens paper.db in SQLite's `mode=ro` URI mode (same pattern as
dashboard.py's _ro_connect) so this tool cannot mutate it. market.db is not
touched at all -- nothing here needs it.
"""
import csv
import os
import sqlite3
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "py"))

import paper_loop as pl
import paper_store as ps

LEDGER_PATH = ROOT / "data" / "trade_ledger.csv"
SUMMARY_PATH = ROOT / "data" / "trade_summary.txt"  # E7 legacy, no longer written
XLSX_PATH = ROOT / "data" / "trade_report.xlsx"
TMP_XLSX_PATH = ROOT / "data" / "trade_report.tmp.xlsx"

RED_FONT = Font(color="FFCC0000")
GREEN_FONT = Font(color="FF008000")
HIGHLIGHT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

OVERVIEW_ROWS = ["已平倉筆數", "勝", "敗", "勝率", "平均獲利", "平均虧損", "賺賠比",
                  "獲利因子", "總損益", "總手續費", "總資金費用", "最大回撤", "目前餘額",
                  "報酬率", "資料起訖"]

LEDGER_HEADER = ["商品", "狀態", "進場時間", "出場時間", "方向", "進場價", "出場價",
                  "手續費", "資金費用", "損益", "損益%", "持有天數", "出場原因"]

ANNOUNCE_LINE1 = "[工程驗證模式] 擠壓突破佔位規則已知無邊際（D-008），本帳僅作觀察，不作策略論證（D-013）"
ANNOUNCE_LINE2 = "實盤門檻（BATCH_PLAN）：IS PF>1.5 且 n>50 → OOS PF>1.2 → 跨兩商品 → CI 下緣扣成本>0"

REASON_LABELS = {"stop": "觸發停損", "reversal": "反手換向", "window_end": "窗末平倉"}


def _ro_connect(path):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"database not found: {p}")
    return sqlite3.connect(f"file:{p.as_posix()}?mode=ro", uri=True)


def fmt_ts(ms):
    if ms is None:
        return ""
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M")


def dirword(direction):
    return "多" if direction == "long" else "空"


def reason_label(reason):
    return REASON_LABELS.get(reason, reason or "")


def fetch_trades(con, symbol):
    cur = con.execute(
        "SELECT entry_ts, exit_ts, direction, qty, entry_px, exit_px, fees, funding, pnl, reason "
        "FROM paper_trades WHERE symbol=? ORDER BY entry_ts", (symbol,))
    cols = ["entry_ts", "exit_ts", "direction", "qty", "entry_px", "exit_px", "fees", "funding",
            "pnl", "reason"]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_state(con, symbol):
    cur = con.execute("SELECT last_ts, cash, equity FROM paper_state WHERE symbol=?", (symbol,))
    row = cur.fetchone()
    return dict(zip(["last_ts", "cash", "equity"], row)) if row else None


def fetch_equity(con, symbol):
    cur = con.execute("SELECT ts, equity FROM paper_equity WHERE symbol=? ORDER BY ts", (symbol,))
    return cur.fetchall()


def _num(v):
    return f"{v:.6f}" if v is not None else ""


def _pct(v):
    return f"{v:.4f}" if v is not None else ""


def build_ledger_records(con, symbols):
    """Per symbol: the open position (if any) first, then closed trades
    newest-to-oldest by exit_ts. Raw-typed (float/None), unlike
    build_ledger_rows()'s pre-formatted strings -- this is the shared source
    for both the CSV rows and the xlsx detail sheet's numeric cells. Pure
    over (con, symbols) -- no filesystem writes here, so it's directly
    unit-testable against a fake sqlite db."""
    records = []
    for symbol in symbols:
        trades = fetch_trades(con, symbol)
        state = fetch_state(con, symbol)
        closed = sorted((t for t in trades if t["exit_ts"] is not None),
                          key=lambda t: t["exit_ts"], reverse=True)
        open_trades = [t for t in trades if t["exit_ts"] is None]

        for t in open_trades:
            cost_basis = t["entry_px"] * t["qty"]
            unreal = (state["equity"] - state["cash"]) if state else None
            pnl_pct = (unreal / cost_basis * 100.0) if (unreal is not None and cost_basis) else None
            hold_days = None
            if state and state.get("last_ts") is not None:
                hold_days = (state["last_ts"] - t["entry_ts"]) / 86400000.0
            records.append(dict(
                symbol=symbol, status="持倉中",
                entry_time=fmt_ts(t["entry_ts"]), exit_time="",
                direction=dirword(t["direction"]),
                entry_px=t["entry_px"], exit_px=None,
                fees=t["fees"], funding=t["funding"], pnl=unreal, pnl_pct=pnl_pct,
                hold_days=hold_days, reason="",
            ))

        for t in closed:
            cost_basis = t["entry_px"] * t["qty"]
            pnl_pct = (t["pnl"] / cost_basis * 100.0) if cost_basis else 0.0
            hold_days = (t["exit_ts"] - t["entry_ts"]) / 86400000.0
            records.append(dict(
                symbol=symbol, status="已平倉",
                entry_time=fmt_ts(t["entry_ts"]), exit_time=fmt_ts(t["exit_ts"]),
                direction=dirword(t["direction"]),
                entry_px=t["entry_px"], exit_px=t["exit_px"],
                fees=t["fees"], funding=t["funding"], pnl=t["pnl"], pnl_pct=pnl_pct,
                hold_days=hold_days, reason=reason_label(t["reason"]),
            ))
    return records


def build_ledger_rows(con, symbols):
    """CSV row order matching LEDGER_HEADER, all pre-formatted strings
    (unchanged E7 CSV contract) -- derived from build_ledger_records()."""
    rows = []
    for r in build_ledger_records(con, symbols):
        rows.append([
            r["symbol"], r["status"], r["entry_time"], r["exit_time"], r["direction"],
            _num(r["entry_px"]), _num(r["exit_px"]), _num(r["fees"]), _num(r.get("funding")),
            _num(r["pnl"]), _pct(r["pnl_pct"]),
            (f'{r["hold_days"]:.2f}' if r["hold_days"] is not None else ""),
            r["reason"],
        ])
    return rows


def write_ledger(rows, path=LEDGER_PATH):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(LEDGER_HEADER)
        w.writerows(rows)


# compute_stats() and max_drawdown_pct() moved verbatim to py/perf.py so the
# dashboard's 策略表現 card and this workbook quote the same numbers from the
# same definitions. Re-exported here: every existing caller (and every test)
# keeps addressing them as export_trades.compute_stats / .max_drawdown_pct.
from perf import compute_stats, max_drawdown_pct  # noqa: E402  (after sys.path setup)


def rolling_max_drawdown_pct(equities):
    """E8 overview-sheet drawdown formula, written verbatim per spec: over a
    ts-ascending list of equity floats, dd_i = equity_i / running_max(equity
    up to i) - 1; result is min(dd_i) as a percentage (e.g. -25.0 for a
    -25% drawdown). Peak seeds at the first observed equity value -- no
    external baseline (unlike max_drawdown_pct()'s cash0-seeded version,
    which format_section()/build_summary() still use unchanged)."""
    if not equities:
        return 0.0
    peak = equities[0]
    min_dd = 0.0
    for eq in equities:
        if eq > peak:
            peak = eq
        dd = (eq / peak - 1.0) if peak else 0.0
        if dd < min_dd:
            min_dd = dd
    return min_dd * 100.0


def _sign_font(value):
    if value is None:
        return None
    if value > 0:
        return GREEN_FONT
    if value < 0:
        return RED_FONT
    return None


def _short_symbol(symbol):
    return symbol.split("/")[0]


def compute_overview(con, symbols, cash0=None):
    """Per-symbol + combined stats/drawdown/balance/return/date-range for the
    xlsx "總覽" sheet. Combined equity is summed by matching ts across
    symbols, then rolling_max_drawdown_pct() is applied to that summed
    series (spec: "合計版先按 ts 加總三商品再同式")."""
    cash0 = pl.CASH0 if cash0 is None else cash0
    per_symbol = {}
    combined_equity_by_ts = {}
    all_closed = []
    combined_last = 0.0
    min_ts_all = max_ts_all = None

    for symbol in symbols:
        trades = fetch_trades(con, symbol)
        closed = [t for t in trades if t["exit_ts"] is not None]
        state = fetch_state(con, symbol)
        equity_rows = fetch_equity(con, symbol)

        stats = compute_stats(closed)
        dd = rolling_max_drawdown_pct([eq for _ts, eq in equity_rows])
        balance = state["equity"] if state else cash0
        ret_pct = (balance / cash0 - 1.0) * 100.0

        if equity_rows:
            date_range = f"{fmt_ts(equity_rows[0][0])} ~ {fmt_ts(equity_rows[-1][0])}"
            min_ts_all = equity_rows[0][0] if min_ts_all is None else min(min_ts_all, equity_rows[0][0])
            max_ts_all = equity_rows[-1][0] if max_ts_all is None else max(max_ts_all, equity_rows[-1][0])
        else:
            date_range = "無資料"

        per_symbol[symbol] = dict(stats=stats, dd=dd, balance=balance,
                                    ret_pct=ret_pct, date_range=date_range)
        all_closed.extend(closed)
        for ts, eq in equity_rows:
            combined_equity_by_ts[ts] = combined_equity_by_ts.get(ts, 0.0) + eq
        combined_last += balance

    combined_stats = compute_stats(all_closed)
    combined_cash0 = cash0 * len(symbols)
    combined_equity_series = [eq for _ts, eq in sorted(combined_equity_by_ts.items())]
    combined_dd = rolling_max_drawdown_pct(combined_equity_series)
    combined_ret = (combined_last / combined_cash0 - 1.0) * 100.0 if combined_cash0 else 0.0
    combined_range = (f"{fmt_ts(min_ts_all)} ~ {fmt_ts(max_ts_all)}"
                       if min_ts_all is not None else "無資料")

    combined = dict(stats=combined_stats, dd=combined_dd, balance=combined_last,
                     ret_pct=combined_ret, date_range=combined_range)
    return per_symbol, combined


def _fill_overview_cell(cell, row_label, data):
    stats = data["stats"]
    if row_label == "已平倉筆數":
        cell.value = stats["n"]
    elif row_label == "勝":
        cell.value = stats["n_win"]
    elif row_label == "敗":
        cell.value = stats["n_loss"]
    elif row_label == "勝率":
        cell.value = stats["win_rate"] / 100.0
        cell.number_format = "0.0%"
    elif row_label == "平均獲利":
        cell.value = round(stats["avg_win"], 2)
        cell.number_format = "#,##0.00"
    elif row_label == "平均虧損":
        cell.value = round(stats["avg_loss"], 2)
        cell.number_format = "#,##0.00"
    elif row_label == "賺賠比":
        if stats["n_loss"] == 0:
            cell.value = "∞（無虧損筆）"
        else:
            cell.value = round(stats["avg_win"] / stats["avg_loss"], 2)
            cell.number_format = "0.00"
    elif row_label == "獲利因子":
        if stats["n_loss"] == 0:
            cell.value = "∞（無虧損筆）"
        else:
            cell.value = round(stats["gross_profit"] / stats["gross_loss"], 2)
            cell.number_format = "0.00"
    elif row_label == "總損益":
        cell.value = round(stats["net"], 2)
        cell.number_format = "#,##0.00"
        font = _sign_font(stats["net"])
        if font:
            cell.font = font
    elif row_label == "總手續費":
        cell.value = round(stats["fees_total"], 2)
        cell.number_format = "#,##0.00"
    elif row_label == "總資金費用":
        cell.value = round(stats["funding_total"], 2)
        cell.number_format = "#,##0.00"
    elif row_label == "最大回撤":
        cell.value = data["dd"] / 100.0
        cell.number_format = "0.00%"
    elif row_label == "目前餘額":
        cell.value = round(data["balance"], 2)
        cell.number_format = "#,##0.00"
    elif row_label == "報酬率":
        cell.value = data["ret_pct"] / 100.0
        cell.number_format = "0.00%"
        font = _sign_font(data["ret_pct"])
        if font:
            cell.font = font
    elif row_label == "資料起訖":
        cell.value = data["date_range"]


def write_overview_sheet(ws, con, symbols, cash0=None):
    cash0 = pl.CASH0 if cash0 is None else cash0
    per_symbol, combined = compute_overview(con, symbols, cash0)

    ws["A1"] = ANNOUNCE_LINE1
    ws["A2"] = ANNOUNCE_LINE2

    header_row = 4
    ws.cell(row=header_row, column=1, value="指標")
    col_symbols = list(symbols) + ["__combined__"]
    col_labels = [_short_symbol(s) for s in symbols] + ["三帳戶合計"]
    for j, label in enumerate(col_labels):
        ws.cell(row=header_row, column=2 + j, value=label)

    for i, row_label in enumerate(OVERVIEW_ROWS):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=row_label)
        for j, sym in enumerate(col_symbols):
            data = combined if sym == "__combined__" else per_symbol[sym]
            cell = ws.cell(row=r, column=2 + j)
            _fill_overview_cell(cell, row_label, data)

    ws.column_dimensions["A"].width = 12
    for j in range(len(col_labels)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16


def write_detail_sheet(ws, records):
    ws.append(LEDGER_HEADER)
    for r in records:
        ws.append([
            r["symbol"], r["status"], r["entry_time"], r["exit_time"], r["direction"],
            r["entry_px"], r["exit_px"], r["fees"], r.get("funding"), r["pnl"], r["pnl_pct"],
            r["hold_days"], r["reason"],
        ])

    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col_letter = get_column_letter(len(LEDGER_HEADER))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    col_idx = {name: i + 1 for i, name in enumerate(LEDGER_HEADER)}
    status_col = col_idx["狀態"]
    entry_px_col = col_idx["進場價"]
    exit_px_col = col_idx["出場價"]
    fees_col = col_idx["手續費"]
    funding_col = col_idx["資金費用"]
    pnl_col = col_idx["損益"]
    pnl_pct_col = col_idx["損益%"]
    hold_days_col = col_idx["持有天數"]
    entry_time_col = col_idx["進場時間"]
    exit_time_col = col_idx["出場時間"]

    for row_i in range(2, last_row + 1):
        status_val = ws.cell(row=row_i, column=status_col).value
        if status_val == "持倉中":
            for c in range(1, len(LEDGER_HEADER) + 1):
                ws.cell(row=row_i, column=c).fill = HIGHLIGHT_FILL

        for col in (entry_px_col, exit_px_col, fees_col, funding_col):
            cell = ws.cell(row=row_i, column=col)
            if cell.value is not None:
                cell.number_format = "#,##0.00"

        pnl_cell = ws.cell(row=row_i, column=pnl_col)
        if pnl_cell.value is not None:
            pnl_cell.number_format = "#,##0.00"
            font = _sign_font(pnl_cell.value)
            if font:
                pnl_cell.font = font

        pnl_pct_cell = ws.cell(row=row_i, column=pnl_pct_col)
        v = pnl_pct_cell.value
        if v is not None:
            pnl_pct_cell.value = v / 100.0
            pnl_pct_cell.number_format = "0.00%"
            font = _sign_font(v)
            if font:
                pnl_pct_cell.font = font

        hold_days_cell = ws.cell(row=row_i, column=hold_days_col)
        if hold_days_cell.value is not None:
            hold_days_cell.number_format = "0.0"

        for col in (entry_time_col, exit_time_col):
            ws.cell(row=row_i, column=col).number_format = "@"

    widths = {"商品": 12, "狀態": 10, "進場時間": 18, "出場時間": 18, "方向": 6,
              "進場價": 14, "出場價": 14, "手續費": 12, "資金費用": 12, "損益": 14,
              "損益%": 10, "持有天數": 10, "出場原因": 14}
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx[name])].width = w


def build_workbook(con, symbols, cash0=None):
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "總覽"
    write_overview_sheet(ws_overview, con, symbols, cash0)

    ws_detail = wb.create_sheet("交易明細")
    records = build_ledger_records(con, symbols)
    write_detail_sheet(ws_detail, records)
    return wb


def write_xlsx(con, symbols, path=None, tmp_path=None, cash0=None):
    """Builds the workbook, saves to a tmp file, then atomically replaces
    the real path via os.replace(). A PermissionError there (the file is
    open in Excel) gets one retry after a 2s wait; if that also fails, print
    the skip line and return normally -- never raises, so a caller like
    main() always exits 0."""
    path = XLSX_PATH if path is None else Path(path)
    tmp_path = TMP_XLSX_PATH if tmp_path is None else Path(tmp_path)
    wb = build_workbook(con, symbols, cash0)
    tmp_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(tmp_path))
    try:
        os.replace(str(tmp_path), str(path))
    except PermissionError:
        time.sleep(2)
        try:
            os.replace(str(tmp_path), str(path))
        except PermissionError:
            print("SKIP: trade_report.xlsx 使用中，本輪跳過")


def format_section(title, stats, max_dd, balance, ret_pct, date_range, cash0):
    no_losses = stats["n_loss"] == 0
    pf = "∞（無虧損筆）" if no_losses else f'{stats["gross_profit"] / stats["gross_loss"]:.2f}'
    wl_ratio = "∞（無虧損筆）" if no_losses else f'{stats["avg_win"] / stats["avg_loss"]:.2f}'
    lines = [
        f"【{title}】",
        f"已平倉筆數：{stats['n']}",
        f"勝/敗：{stats['n_win']}/{stats['n_loss']}",
        f"勝率：{stats['win_rate']:.1f}%",
        f"平均獲利：{stats['avg_win']:,.2f}",
        f"平均虧損：{stats['avg_loss']:,.2f}",
        f"賺賠比：{wl_ratio}",
        f"獲利因子：{pf}",
        f"總損益：{stats['net']:+,.2f}",
        f"最大回撤：{max_dd:.2f}%",
        f"目前餘額：{balance:,.2f}",
        f"報酬率（基準 {cash0:,.0f}）：{ret_pct:+.2f}%",
        f"資料起訖：{date_range}",
    ]
    return "\n".join(lines)


def build_summary(con, symbols, cash0=None):
    cash0 = pl.CASH0 if cash0 is None else cash0
    sections = []
    combined_equity_by_ts = {}
    all_closed = []
    combined_last = 0.0
    min_ts_all = None
    max_ts_all = None

    for symbol in symbols:
        trades = fetch_trades(con, symbol)
        closed = [t for t in trades if t["exit_ts"] is not None]
        state = fetch_state(con, symbol)
        equity_rows = fetch_equity(con, symbol)

        stats = compute_stats(closed)
        max_dd = max_drawdown_pct(equity_rows, cash0)
        balance = state["equity"] if state else cash0
        ret_pct = (balance / cash0 - 1.0) * 100.0

        if equity_rows:
            t0, t1 = equity_rows[0][0], equity_rows[-1][0]
            date_range = f"{fmt_ts(t0)} ~ {fmt_ts(t1)}"
            min_ts_all = t0 if min_ts_all is None else min(min_ts_all, t0)
            max_ts_all = t1 if max_ts_all is None else max(max_ts_all, t1)
        else:
            date_range = "無資料"

        sections.append(format_section(symbol, stats, max_dd, balance, ret_pct, date_range, cash0))

        all_closed.extend(closed)
        for ts, eq in equity_rows:
            combined_equity_by_ts[ts] = combined_equity_by_ts.get(ts, 0.0) + eq
        combined_last += balance

    combined_stats = compute_stats(all_closed)
    combined_cash0 = cash0 * len(symbols)
    combined_equity_rows = sorted(combined_equity_by_ts.items())
    combined_dd = max_drawdown_pct(combined_equity_rows, combined_cash0)
    combined_ret = (combined_last / combined_cash0 - 1.0) * 100.0 if combined_cash0 else 0.0
    combined_range = (f"{fmt_ts(min_ts_all)} ~ {fmt_ts(max_ts_all)}"
                       if min_ts_all is not None else "無資料")

    sections.append(format_section("三帳戶合計", combined_stats, combined_dd, combined_last,
                                    combined_ret, combined_range, combined_cash0))
    return "\n\n".join(sections)


def write_summary(con, symbols, path=SUMMARY_PATH):
    body = build_summary(con, symbols)
    text = ANNOUNCE_LINE1 + "\n" + ANNOUNCE_LINE2 + "\n\n" + body + "\n"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def main():
    con = _ro_connect(ps.DB_PATH)
    try:
        rows = build_ledger_rows(con, pl.SYMBOLS)
        write_ledger(rows)
        write_xlsx(con, pl.SYMBOLS)
    finally:
        con.close()
    print(f"wrote {LEDGER_PATH} ({len(rows)} rows)")
    print(f"wrote {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
