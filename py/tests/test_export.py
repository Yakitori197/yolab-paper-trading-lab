"""E7/E8 tests for tools/export_trades.py. Uses a temporary sqlite db built
from paper_store.py's own schema and hand-crafted rows -- never touches the
real data/paper.db."""
import sqlite3
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import export_trades as et
import paper_store as ps


def make_db(tmp_path):
    path = tmp_path / "fake_paper.db"
    con = sqlite3.connect(str(path))
    con.executescript(ps.SCHEMA)
    return con


def row_by_status(rows, status):
    header = et.LEDGER_HEADER
    return [dict(zip(header, r)) for r in rows if r[header.index("狀態")] == status][0]


def test_ledger_rows_pct_days_direction_status_and_unrealized(tmp_path):
    con = make_db(tmp_path)
    symbol = "TEST/USDT"
    day = 86400000

    # closed long trade: entry T0, exit T0+2d, pnl=19.0, cost basis 100*2=200 -> pnl% = 9.5
    t0 = 1_700_000_000_000
    con.execute(
        "INSERT INTO paper_trades (symbol, entry_ts, exit_ts, direction, qty, entry_px, exit_px, "
        "fees, pnl, reason) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (symbol, t0, t0 + 2 * day, "long", 2.0, 100.0, 110.0, 1.0, 19.0, "stop"))

    # open short trade: entry T2, no exit yet
    t2 = 1_700_500_000_000
    con.execute(
        "INSERT INTO paper_trades (symbol, entry_ts, exit_ts, direction, qty, entry_px, exit_px, "
        "fees, pnl, reason) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (symbol, t2, None, "short", 1.5, 50.0, None, 0.075, None, None))

    # paper_state: last_ts = t2 + 1d (hold_days=1.0), cash/equity gives unrealized = 25.0
    con.execute(
        "INSERT INTO paper_state (symbol, last_ts, cash, position_dir, qty, entry_px, entry_ts, "
        "stop_disp, equity, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (symbol, t2 + day, 9950.0, "short", 1.5, 50.0, t2, None, 9975.0, "2026-01-01T00:00:00Z"))
    con.commit()

    rows = et.build_ledger_rows(con, [symbol])
    assert len(rows) == 2

    closed = row_by_status(rows, "已平倉")
    assert closed["方向"] == "多"
    assert closed["持有天數"] == "2.00"
    assert float(closed["損益%"]) == pytest.approx(9.5, abs=1e-6)

    open_row = row_by_status(rows, "持倉中")
    assert open_row["方向"] == "空"
    assert open_row["持有天數"] == "1.00"
    assert float(open_row["損益"]) == pytest.approx(25.0, abs=1e-6)  # equity - cash
    assert float(open_row["損益%"]) == pytest.approx(25.0 / (50.0 * 1.5) * 100.0, abs=1e-4)
    assert open_row["出場時間"] == ""
    assert open_row["出場價"] == ""

    con.close()


def test_summary_stats_match_hand_calc_for_3_wins_2_losses(tmp_path):
    con = make_db(tmp_path)
    symbol = "STAT/USDT"
    t0 = 1_700_000_000_000
    day = 86400000

    # 3 wins: 10, 20, 30 -> gross_profit=60, avg_win=20
    # 2 losses: -5, -15 -> gross_loss=20, avg_loss=10
    pnls = [10.0, 20.0, 30.0, -5.0, -15.0]
    for i, pnl in enumerate(pnls):
        ts = t0 + i * day
        con.execute(
            "INSERT INTO paper_trades (symbol, entry_ts, exit_ts, direction, qty, entry_px, exit_px, "
            "fees, pnl, reason) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (symbol, ts, ts + day, "long", 1.0, 100.0, 100.0 + pnl, 0.1, pnl, "stop"))
    con.commit()

    closed = [t for t in et.fetch_trades(con, symbol) if t["exit_ts"] is not None]
    stats = et.compute_stats(closed)

    assert stats["n"] == 5
    assert stats["n_win"] == 3
    assert stats["n_loss"] == 2
    assert stats["win_rate"] == pytest.approx(60.0)
    assert stats["avg_win"] == pytest.approx(20.0)
    assert stats["avg_loss"] == pytest.approx(10.0)
    wl_ratio = stats["avg_win"] / stats["avg_loss"]
    pf = stats["gross_profit"] / stats["gross_loss"]
    assert wl_ratio == pytest.approx(2.0)
    assert pf == pytest.approx(3.0)

    con.close()


def test_overview_max_drawdown_stored_as_fraction_with_pct_format():
    # hand calc: dd = -25.0 (percentage units) -> stored fraction -0.25, format 0.00%
    data = dict(stats=et.compute_stats([]), dd=-25.0, balance=0.0, ret_pct=0.0, date_range="")
    ws = Workbook().active
    cell = ws.cell(row=1, column=1)
    et._fill_overview_cell(cell, "最大回撤", data)
    assert cell.value == pytest.approx(-0.25)
    assert cell.number_format == "0.00%"


def test_detail_pnl_pct_stored_as_fraction_with_pct_format():
    records = [dict(
        symbol="TEST/USDT", status="已平倉",
        entry_time="2025-01-01 00:00", exit_time="2025-01-02 00:00",
        direction="多", entry_px=100.0, exit_px=110.0,
        fees=1.0, pnl=19.0, pnl_pct=9.5,
        hold_days=2.0, reason="觸發停損",
    )]
    ws = Workbook().create_sheet("交易明細")
    et.write_detail_sheet(ws, records)
    col_idx = {name: i + 1 for i, name in enumerate(et.LEDGER_HEADER)}

    pnl_pct_cell = ws.cell(row=2, column=col_idx["損益%"])
    assert pnl_pct_cell.value == pytest.approx(0.095)
    assert pnl_pct_cell.number_format == "0.00%"

    fees_cell = ws.cell(row=2, column=col_idx["手續費"])
    assert fees_cell.number_format == "#,##0.00"

    hold_days_cell = ws.cell(row=2, column=col_idx["持有天數"])
    assert hold_days_cell.number_format == "0.0"


def test_ledger_csv_has_utf8_sig_bom_and_full_header(tmp_path):
    con = make_db(tmp_path)
    symbol = "CSV/USDT"
    t0 = 1_700_000_000_000
    day = 86400000
    con.execute(
        "INSERT INTO paper_trades (symbol, entry_ts, exit_ts, direction, qty, entry_px, exit_px, "
        "fees, pnl, reason) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (symbol, t0, t0 + day, "long", 1.0, 100.0, 105.0, 0.1, 4.9, "stop"))
    con.commit()

    rows = et.build_ledger_rows(con, [symbol])
    out_path = tmp_path / "ledger.csv"
    et.write_ledger(rows, path=out_path)

    raw = out_path.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")

    text = raw.decode("utf-8-sig")
    header_line = text.splitlines()[0]
    assert header_line.split(",") == et.LEDGER_HEADER

    con.close()
